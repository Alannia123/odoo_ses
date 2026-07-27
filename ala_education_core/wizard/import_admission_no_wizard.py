import base64
import io
import openpyxl

from odoo import models, fields, _
from odoo.exceptions import UserError


class ImportAdmissionNoWizard(models.TransientModel):
    _name = 'import.admission.no.wizard'
    _description = 'Import Admission Number'

    file = fields.Binary(string="Excel File", required=True)
    file_name = fields.Char(string="File Name")

    def action_import_admission_no(self):
        if not self.file:
            raise UserError(_("Please upload Excel file."))

        def format_admission_no(value):
            value = str(value or '').strip()
            if '/' in value:
                left, right = value.rsplit('/', 1)
                right = right.strip()
                if len(right) == 2 and right.isdigit():
                    value = "%s/20%s" % (left.strip(), right)
            return value

        try:
            data = base64.b64decode(self.file)
            workbook = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
            sheet = workbook.active
        except Exception as e:
            raise UserError(_("Invalid Excel file: %s") % e)

        headers = {}
        for col in range(1, sheet.max_column + 1):
            value = sheet.cell(row=1, column=col).value
            if value:
                headers[str(value).strip().lower()] = col

        required_cols = ['name', 'registration number', 'admission no']
        for col in required_cols:
            if col not in headers:
                raise UserError(_("Missing column in Excel: %s") % col)

        Student = self.env['ala.education.student']

        updated = 0
        not_found = []

        for row in range(2, sheet.max_row + 1):
            name = sheet.cell(row=row, column=headers['name']).value
            registration_no = sheet.cell(row=row, column=headers['registration number']).value
            admission_no = sheet.cell(row=row, column=headers['admission no']).value

            if not name or not registration_no or not admission_no:
                continue

            name = str(name).strip()
            registration_no = str(registration_no).strip()
            admission_no = format_admission_no(admission_no)

            students = Student.search([
                ('register_no', '=', registration_no),
            ])

            if not students:
                not_found.append("Row %s: Name not found - %s" % (row, name))
                continue

            student = students.filtered(
                lambda s: str(s.register_no or '').strip() == registration_no
            )

            if not student:
                not_found.append(
                    "Row %s: Registration number not matched for %s" % (row, name)
                )
                continue

            student_rec = student[0].sudo()

            print("BEFORE:", student_rec.id, student_rec.admission_no)

            student_rec.write({
                'admission_no': admission_no
            })

            self.env.cr.commit()

            student_rec.invalidate_recordset(['admission_no'])

            print("AFTER WRITE:", student_rec.admission_no)

            self.env.cr.execute("""
                SELECT admission_no 
                FROM ala_education_student 
                WHERE id = %s
            """, [student_rec.id])

            print("DB VALUE:", self.env.cr.fetchone())

            updated += 1

        message = "Admission No updated successfully: %s" % updated

        if not_found:
            message += "\n\nNot matched:\n" + "\n".join(not_found[:50])

        raise UserError(_(message))